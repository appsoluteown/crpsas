#!/usr/bin/env python3
"""
Génère public/data.json à partir de l'Excel "Récap transports tarifs".

Onglets pris en compte : "Général - Taut" et "Général - plateau"
(ce sont les seules grilles affichées par l'application).

Usage:
    python3 build_data.py "/chemin/vers/Recap transports tarifs.xlsx"

Le script reproduit fidèlement le schéma JSON existant :
  - Taut    : prix TAXE GAZOLE COMPRISE (colonnes J..N), arrondis 2 décimales,
              cellule vide -> 0.0, avec facteurs_taxe_gazole (calculés base->gazole).
  - Plateau : prix bruts hors gazole (colonnes C..E).
  - transporteurs triés alphabétiquement, prix triés par valeur croissante.
"""
import sys
import json
import datetime
import openpyxl

# Configuration des deux feuilles Général ------------------------------------
SHEETS = [
    {
        "sheet": "Général - Taut",
        "nom": "Général - Taut",
        "type": "Taut",
        "info": "TAXE GAZOLE COMPRISE",
        "header_row": 10,
        "dept_col": 3,  # C
        # Prix TAXE GAZOLE COMPRISE (bloc de droite)
        "price_cols": {"GEODIS": 10, "KUEHNE": 11, "LACHAUD24": 12, "XPO": 13, "SCHENKER": 14},  # J..N
        # Prix de base HT (bloc de gauche) -> sert à calculer les facteurs gazole
        "base_cols": {"GEODIS": 4, "KUEHNE": 5, "LACHAUD24": 6, "XPO": 7, "SCHENKER": 8},  # D..H
        "gazole": True,
    },
    {
        "sheet": "Général - plateau",
        "nom": "Général - plateau",
        "type": "Plateau",
        "info": "Hors taxe Gazole",
        "header_row": 10,
        "dept_col": 2,  # B
        "price_cols": {"CONTE": 3, "VEYSSET": 4, "LACHAUD": 5},  # C..E
        "gazole": False,
    },
]


def num(v):
    """Arrondit à 2 décimales en conservant int pour les entiers (style du JSON existant)."""
    if v is None:
        return None
    r = round(float(v), 2)
    if r == int(r):
        return int(r)
    return r


def build_sheet(ws, cfg):
    hr = cfg["header_row"]

    # Validation : les en-têtes de colonnes doivent correspondre aux transporteurs attendus
    for carrier, col in cfg["price_cols"].items():
        got = ws.cell(row=hr, column=col).value
        assert got == carrier, (
            f"[{cfg['sheet']}] En-tête colonne {col} = {got!r}, attendu {carrier!r}. "
            f"La structure de l'Excel a changé — vérifier le mapping des colonnes."
        )

    tarifs = []
    r = hr + 1
    while True:
        dept = ws.cell(row=r, column=cfg["dept_col"]).value
        if dept is None or str(dept).strip() == "":
            break  # fin des données
        dept = str(dept).strip()

        prix = {}
        for carrier, col in cfg["price_cols"].items():
            raw = ws.cell(row=r, column=col).value
            if cfg["gazole"]:
                prix[carrier] = num(raw) if raw is not None else 0.0
            else:
                prix[carrier] = num(raw) if raw is not None else 0

        # Tri par prix croissant (style du JSON existant)
        prix_tri = dict(sorted(prix.items(), key=lambda kv: (kv[1], kv[0])))
        tarifs.append({"departement": dept, "prix_par_transporteur": prix_tri})
        r += 1

    feuille = {
        "nom": cfg["nom"],
        "type": cfg["type"],
        "info": cfg["info"],
    }

    if cfg["gazole"]:
        # Facteur gazole par transporteur = prix_gazole / prix_base (1er échantillon valide)
        facteurs = {}
        for carrier in cfg["price_cols"]:
            gcol, bcol = cfg["price_cols"][carrier], cfg["base_cols"][carrier]
            for rr in range(hr + 1, r):
                base = ws.cell(row=rr, column=bcol).value
                gaz = ws.cell(row=rr, column=gcol).value
                if base and gaz:  # tous deux non nuls / non vides
                    facteurs[carrier] = round(gaz / base, 3)
                    break
        feuille["facteurs_taxe_gazole"] = dict(sorted(facteurs.items()))

    feuille["pays"] = "France"
    feuille["transporteurs"] = sorted(cfg["price_cols"].keys())
    feuille["nombre_departements"] = len(tarifs)
    feuille["tarifs"] = tarifs
    return feuille


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 build_data.py <chemin_excel.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    feuilles = []
    for cfg in SHEETS:
        ws = wb[cfg["sheet"]]
        feuille = build_sheet(ws, cfg)
        feuilles.append(feuille)
        print(f"  ✓ {cfg['nom']:20} → {feuille['nombre_departements']} départements, "
              f"{len(feuille['transporteurs'])} transporteurs")

    data = {
        "fichier": xlsx_path.split("/")[-1],
        "description": "Grille tarifaire transport - France",
        "date_extraction": datetime.date.today().isoformat(),
        "feuilles": feuilles,
    }

    out_path = "public/data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"\n→ {out_path} généré ({data['date_extraction']}, fichier source: {data['fichier']})")


if __name__ == "__main__":
    main()
